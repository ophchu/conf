from datetime import datetime
from openpyxl import load_workbook, Workbook
import csv


def validate(date_text):
    try:
        datetime.strptime(date_text, '%d/%m/%Y')
        return True
    except ValueError:
        return False


class Transaction:
    def __init__(self, month_year, date, name, amount, account, cat_lvl1: str, cat_lvl2: str):
        self.month_year = month_year
        self.date = date
        self.name = name
        self.amount = amount
        self.account = account
        self.cat_level1: str = cat_lvl1
        self.cat_level2: str = cat_lvl2

    def __iter__(self):
        return iter(
            [self.month_year, self.date, self.name, self.amount, self.account, self.cat_level1, self.cat_level2])

    def __str__(self) -> str:
        return """my: [%s], date: [%s], desc: [%s], amount: [%.2f], account: [%s]. Category: [%s, %s]""" % \
               (self.month_year, self.date, self.name, self.amount, self.account, self.cat_level1, self.cat_level2)


def write_transactions(transactions, output):
    header = ['month', 'date', 'name', 'amount', 'account', 'cat1', 'cat2']
    wb = Workbook()

    sheet = wb.create_sheet("total")
    sheet.append(header)
    for txs in transactions:
        sheet.append((txs.month_year, txs.date, txs.name, txs.amount, txs.account, txs.cat_level1, txs.cat_level2))

    wb.save(output)


def read_isracard(file, cat_match: dict):
    month = '_'.join(file.split('/')[-1].split('_')[1:3])
    workbook = load_workbook(filename=file)
    sheet = workbook["Transactions"]
    transaction = []
    account = ''
    abroad = False

    for row in sheet.iter_rows():
        if row[0].value is not None:
            if row[1].value == 'מועד חיוב':
                account = row[0].value.replace('*', '').split('-')[-1].strip()
                abroad = False
            elif row[0].value == 'עסקאות בחו˝ל':
                abroad = True
            elif validate(row[0].value) is True:
                if abroad:
                    cat1 = ''
                    cat2 = ''
                    if row[2].value in cat_match:
                        cat1 = cat_match[row[2].value][0]
                        cat2 = cat_match[row[2].value][1]
                    transaction.append(Transaction(month, row[0].value, row[2].value, row[5].value, account,
                                                         cat1, cat2))
                else:
                    cat1 = ''
                    cat2 = ''
                    if row[1].value in cat_match:
                        cat1 = cat_match[row[1].value][0]
                        cat2 = cat_match[row[1].value][1]
                    transaction.append(Transaction(month, row[0].value, row[1].value, row[4].value, account,
                                                         cat1, cat2))
    return transaction


def write_outputfile(transactions, output_file):
    header = ['month', 'date', 'name', 'amount', 'account', 'cat1', 'cat2']
    with open(output_file, 'w', encoding='UTF8') as f:
        # create the csv writer
        writer = csv.writer(f)

        writer.writerow(header)
        writer.writerows(transactions)
